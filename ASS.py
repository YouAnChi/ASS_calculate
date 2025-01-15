from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
import numpy as np
import time
#模型下载



def seconds_to_hms(seconds):
    seconds = int(seconds)
    # 转换小时
    hours = seconds // 3600
    # 计算剩余秒数
    seconds %= 3600
    # 转换分钟
    minutes = seconds // 60
    # 计算剩余秒数
    seconds %= 60
    # 返回格式化的字符串
    return f"总共用时：{hours:02d}:{minutes:02d}:{seconds:02d}"


def ragas(reference_answer, generated_answer):
    # 计算两个答案的嵌入向量
    embedding_1 = np.array(model.encode(reference_answer))
    embedding_2 = np.array(model.encode(generated_answer))

    # 计算余弦相似度
    norms_1 = np.linalg.norm(embedding_1, keepdims=True)
    norms_2 = np.linalg.norm(embedding_2, keepdims=True)
    embedding_1_normalized = embedding_1 / norms_1
    embedding_2_normalized = embedding_2 / norms_2
    similarity = embedding_1_normalized @ embedding_2_normalized.T
    score = similarity.flatten()
    similarity_score = score.tolist()[0]

    return similarity_score


def get_column_data(column_name):
    # 找到列标题对应的列索引
    column_index = None
    for col_idx, cell in enumerate(ws[1], start=1):  # 假设第一行是标题行
        if cell.value == column_name:
            column_index = col_idx
            break

    if column_index is None:
        print(f"列名 '{column_name}' 不存在于Excel文件的标题行中。")
        exit()

    # 读取整列的内容
    column_data = [ws.cell(row=row, column=column_index).value for row in range(2, ws.max_row + 1)]  # 假设从第二行开始是数据
    return column_data


def get_column_index(column_name):
    # 找到列标题对应的列索引
    column_index = None
    for col_idx, cell in enumerate(ws[1], start=1):  # 假设第一行是标题行
        if cell.value == column_name:
            column_index = col_idx
            break

    if column_index is None:
        print(f"列名 '{column_name}' 不存在于Excel文件的标题行中。")
        exit()
    return column_index


def assessz(file_path, ref, ans, data):
    # 替换为你的Excel文件路径
    similarity_column_index = get_column_index(data)
    data_to_write = []
    reference = get_column_data(ref)
    answer = get_column_data(ans)
    for index, (item1, item2) in enumerate(zip(reference, answer), start=1):
        if item1 is None:
            break
        print(f"第{index}行相似度计算")
        similarity = ragas(item1, item2)
        data_to_write.append(similarity)

    # 写入数据，从第二行开始（假设第一行是标题行）

    for row_idx, value in enumerate(data_to_write, start=2):  # 从第二行开始，行索引从2开始
        ws.cell(row=row_idx, column=similarity_column_index, value=value)

    # 保存工作簿
    wb.save(file_path)


def assess(file_path, ref, ans, data):
    # 替换为你的Excel文件路径
    question_idxs = 0
    similarity_column_index = get_column_index(data)
    data_to_write = []
    reference = get_column_data(ref)
    answer = get_column_data(ans)
    row_count = 0
    for index, (item1, item2) in enumerate(zip(reference, answer), start=1):
        if item1 is None:
            break
        print(f"第{index}行相似度计算")
        try:
            similarity = ragas(item1, item2)
            data_to_write.append(similarity)
            row_count += 1
            if row_count % batch_size == 0:  # 每50行写入一次并保存
                for row_idx, value in enumerate(data_to_write, start=index + 2 - batch_size):
                    ws.cell(row=row_idx, column=similarity_column_index, value=value)
                wb.save(file_path)
                data_to_write = []  # 重置列表，准备下一批写入
                question_idxs = index
        except Exception as e:
            print(f"处理第{index}行时发生错误: {e}")

            # 处理剩余数据（如果有）
    if data_to_write:
        for row_idx, value in enumerate(data_to_write, start=question_idxs + 2):
            ws.cell(row=row_idx, column=similarity_column_index, value=value)
        wb.save(file_path)


if __name__ == '__main__':
    # 记录开始时间
    start_time = time.time()

    model = SentenceTransformer(r'/Users/lpd/Documents/acge_text_embedding/modelscope/hub/yangjhchs/acge_text_embedding')
    # excel文件路径
    file_path = '123.xlsx'
    # 对比数据的表头
    ref = '标注'
    ans = 'answer'
    # 输出相似度的表头
    data = 'ASS'
    batch_size = 50  # 每批处理的问题数量
    wb = load_workbook(file_path)
    ws = wb.active
    assess(file_path, ref, ans, data)

    # 记录结束时间并计算运行时间
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(elapsed_time)
    elapsed = seconds_to_hms(elapsed_time)
    print(elapsed)
